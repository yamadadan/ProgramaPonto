from tkinter import font, Label, Toplevel, filedialog, messagebox
from datetime import datetime
import os
import calendar
from openpyxl import Workbook, load_workbook

def iniciar_arrastar(event):
  global x, y
  x = event.x
  y = event.y

def arrastar_janela(event, janela):
  x_deslocamento = event.x - x
  y_deslocamento = event.y - y
  janela.geometry(f"+{janela.winfo_x() + x_deslocamento}+{janela.winfo_y() + y_deslocamento}")

def modificar_diretorio():
  diretorio = carregar_diretorio_salvo()
  diretorio = filedialog.askdirectory(title='Por favor, selecione um diretório para salvar os horários', initialdir=diretorio)
  if diretorio:
    # Se o usuário escolheu um diretório, salve-o para uso futuro
    with open("diretorio_salvo.txt", "w") as arquivo:
      arquivo.write(diretorio)
    messagebox.showinfo("Sucesso", f"Diretório selecionado: {diretorio}")
  # return diretorio

def selecionar_diretorio():
  diretorio = filedialog.askdirectory(title='Por favor, selecione um diretório para salvar os horários')
  if diretorio:
    # Se o usuário escolheu um diretório, salve-o para uso futuro
    with open("diretorio_salvo.txt", "w") as arquivo:
      arquivo.write(diretorio)
    messagebox.showinfo("Sucesso", f"Diretório selecionado: {diretorio}")
  return diretorio

def carregar_diretorio_salvo():
  try:
    # Tente carregar o diretório salvo da última execução
    with open("diretorio_salvo.txt", "r") as arquivo:
      diretorio_salvo = arquivo.read()
      return diretorio_salvo
  except FileNotFoundError:
    return None
    
def mostrar_mensagem(janela):
  # Cria uma nova janela sem decoração
  mensagem_janela = Toplevel(janela)
  mensagem_janela.overrideredirect(True)

  # Adiciona o horário atual na nova janela
  horario_atual = datetime.now().strftime("%H:%M:%S")
  estilo_fonte = font.Font(family="Helvetica", size=12, weight="bold")
  mensagem_label = Label(mensagem_janela, text=f"Sucesso!\nHorário de {horario_atual} adicionado na planilha!", font=estilo_fonte)
  mensagem_label.pack(padx=70, pady=50)

  # Move a nova janela para o topo da janela principal
  x_pos = janela.winfo_x() + (janela.winfo_width() - mensagem_janela.winfo_reqwidth()) // 2
  y_pos = janela.winfo_y() + (janela.winfo_height() - mensagem_janela.winfo_reqheight()) // 2
  mensagem_janela.geometry(f"+{x_pos}+{y_pos}")
  # Move a nova janela para o topo
  # mensagem_janela.lift(janela)
  # Move a nova janela para o topo
  # mensagem_janela.attributes('-topmost', True)

  # Aguarda 3000 milissegundos (3 segundos) e fecha a nova janela
  mensagem_janela.after(1500, mensagem_janela.destroy)

def marcar_ponto(janela, nome_arquivo):
  # Obter o horário atual
  horario_atual = datetime.now().strftime('%H:%M:%S')   
  dia_atual = datetime.now().strftime('%Y-%m-%d')

  # # Nome do arquivo da planilha
  # nome_arquivo = 'ponto.xlsx'

  # Se o arquivo da planilha não existir, crie uma nova planilha
  if not os.path.exists(nome_arquivo):
    planilha = Workbook()
    sheet = planilha.active
    # Preencher as células A1, B1 e C1
    sheet['A1'] = 'Data'
    sheet['B1'] = 'Horario Entrada 1'
    sheet['C1'] = 'Horario Saída 1'
    sheet['D1'] = 'Horario Entrada 2'
    sheet['E1'] = 'Horario Saída 2'
    sheet['F1'] = 'Horas Trabalhadas'
    sheet['G1'] = 'Horas Faltantes'
    planilha.save(nome_arquivo)

  # Adicionar o horário atual à planilha Excel
  planilha = load_workbook(nome_arquivo)
  # workbook = openpyxl.load_workbook(caminho_arquivo)

  # Obter a lista de abas
  abas = planilha.sheetnames

  # Selecionar a última aba
  folha = planilha[abas[-1]]
  # folha = planilha.active

  # Criar uma lista para armazenar os valores da coluna A
  valores_Data = []
  valores_Entrada1 = []
  valores_Saida1 = []
  valores_Entrada2 = []
  valores_Saida2 = []

  # Iterar pelas linhas e obter os valores das colunas
  for row in folha.iter_rows(min_row=2, max_row=folha.max_row, min_col=1, max_col=folha.max_column):
    for cell in row:
      if cell.column_letter == 'A':
        if cell.value is not None:
          valores_Data.append(cell.value)
      elif cell.column_letter == 'B':
        if cell.value is not None:
          valores_Entrada1.append(cell.value)
      elif cell.column_letter == 'C':
        if cell.value is not None:
          valores_Saida1.append(cell.value)
      elif cell.column_letter == 'D':
        if cell.value is not None:
          valores_Entrada2.append(cell.value)
      elif cell.column_letter == 'E':
        if cell.value is not None:
          valores_Saida2.append(cell.value)
  
  # Se temos valores de data
  if(len(valores_Data)):
    if datetime.strptime(valores_Data[-1], "%Y-%m-%d").month == datetime.now().month:
      # Se ultima data for menor que a data atual, inclui valores
      if valores_Data[-1] < dia_atual:
        i = str(len(valores_Data) + 2)
        folha['A'+i] = dia_atual
        folha['B'+i] = horario_atual
        folha['F'+i] = f"=C{i} - B{i} + (E{i} - D{i})"
        folha['F'+i].number_format = 'hh:mm:ss'
        folha['G'+i] = f"=8 - F{i} + F{i-1}"
        folha['G'+i].number_format = 'hh:mm:ss'

      elif valores_Data[-1] == dia_atual:
        if len(valores_Data) > len(valores_Entrada1):
          folha['B'+str(len(valores_Entrada1) + 2)] = horario_atual
        elif len(valores_Data) > len(valores_Saida1):
          folha['C'+str(len(valores_Saida1) + 2)] = horario_atual
        elif len(valores_Data) > len(valores_Entrada2):
          folha['D'+str(len(valores_Entrada2) + 2)] = horario_atual
        elif len(valores_Data) > len(valores_Saida2):
          folha['E'+str(len(valores_Saida2) + 2)] = horario_atual
    else:
      planilha = load_workbook(nome_arquivo)
      # folha = planilha.active
      # Adicione uma nova aba
      sheet = planilha.create_sheet(calendar.month_abbr[datetime.now().month])
      sheet['A1'] = 'Data'
      sheet['B1'] = 'Horario Entrada 1'
      sheet['C1'] = 'Horario Saída 1'
      sheet['D1'] = 'Horario Entrada 2'
      sheet['E1'] = 'Horario Saída 2'
      sheet['F1'] = 'Horas Trabalhadas'
      sheet['G1'] = 'Horas Faltantes'
      # Salve o arquivo
      planilha.save(nome_arquivo)
  # Caso ainda não tenhamos valores de data
  else:
    i = str(len(valores_Data) + 2)
    folha['A'+i] = dia_atual
    folha['B'+i] = horario_atual
    folha['F'+i] = f"=C{i} - B{i} + (E{i} - D{i})"
    folha['F'+i].number_format = 'hh:mm:ss'
    folha['G'+i] = f"=8 - F{i}"
    folha['G'+i].number_format = 'hh:mm:ss'
  
  # Ajustar largura das colunas
  for column in folha.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
      try:
        if len(str(cell.value)) > max_length:
          max_length = len(cell.value)
      except:
        pass
    adjusted_width = (max_length + 2)
    folha.column_dimensions[column[0].column_letter].width = adjusted_width

  # folha.append([horario_atual])
  planilha.save(nome_arquivo)

  # Exibe uma mensagem
  # messagebox.showinfo("Sucesso", "Horário adicionado na planilha!")
  mostrar_mensagem(janela)
  
  # Aguarda 3000 milissegundos (3 segundos) e fecha a janela
  janela.after(1500, janela.destroy)